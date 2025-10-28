# -*- coding: utf-8 -*-
"""
VLMFileSummarizer — AgentLego tool
----------------------------------

A custom AgentLego tool that summarizes local data files (JSON, CSV, TXT, XLSX, YAML)
into a compact, VLM-friendly text brief. Useful as a preprocessing step before sending
context into a Vision/Multimodal LLM.

Features
- Accepts one or many file paths, directories, or globs.
- Supports: .json, .csv, .txt, .xlsx, .yml, .yaml
- Produces concise textual profiles (shape, schema, stats, key-value outlines, samples).
- Honors size/row/sample caps to keep output within model context.
- Safe fallbacks when optional libraries are absent.

Usage
-----
# 1) Programmatic (pure Python)
from vlm_file_summarizer import VLMFileSummarizer
tool = VLMFileSummarizer()
print(tool(['data/*.csv', 'notes.md'], max_output_chars=8000))

# 2) LangChain
from langchain import hub
from langchain.agents import create_structured_chat_agent, AgentExecutor
from langchain_openai import ChatOpenAI
tool = VLMFileSummarizer()
agent = create_structured_chat_agent(
    llm=ChatOpenAI(temperature=0.),
    tools=[tool.to_langchain()],
    prompt=hub.pull("hwchase17/structured-chat-agent"),
)
agent_exec = AgentExecutor(agent=agent, tools=[tool.to_langchain()], verbose=True)
agent_exec.invoke({"input": "Summarize ./data/*.csv for me"})

# 3) Lagent (ReAct)
from lagent import ReAct, GPTAPI, ActionExecutor
tool = VLMFileSummarizer()
agent = ReAct(llm=GPTAPI(temperature=0.), action_executor=ActionExecutor([tool.to_lagent()]))
print(agent.chat("Summarize ./data/sales_2024.xlsx").response)

# 4) Tool server
agentlego-server start --extra ./vlm_file_summarizer.py VLMFileSummarizer

Requirements
------------
- Python 3.8+
- Optional (recommended): pandas, pyyaml, openpyxl

Install extras:
pip install pandas pyyaml openpyxl

License
-------
MIT
"""

from __future__ import annotations

import csv
import io
import json
import math
import os
import re
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable, List, Optional, Sequence, Tuple, Union, Dict, Any

# AgentLego base
from agentlego.tools import BaseTool

# Optional dependencies (lazily bound in setup())
_pd = None          # pandas
_yaml = None        # pyyaml
_openpyxl = None    # openpyxl


# ------------------------------- Utilities --------------------------------- #

SUPPORTED_EXTS = {".json", ".csv", ".txt", ".xlsx", ".yml", ".yaml", ".md"}

def _human_bytes(n: int) -> str:
    if n is None:
        return "unknown"
    units = ["B", "KB", "MB", "GB", "TB"]
    if n <= 0:
        return "0 B"
    i = min(int(math.log(n, 1024)), len(units) - 1)
    return f"{n / (1024 ** i):.2f} {units[i]}"

def _safe_read_text(path: Path, max_bytes: int = 200_000, encoding_guess: str = "utf-8") -> Tuple[str, str]:
    """
    Returns (text_preview, encoding_used). Reads up to max_bytes from file.
    Tries UTF-8 first, falls back to latin-1.
    """
    size = path.stat().st_size if path.exists() else 0
    to_read = min(size, max_bytes) if size else max_bytes
    with open(path, "rb") as f:
        raw = f.read(to_read)
    for enc in [encoding_guess, "utf-8", "utf-16", "latin-1"]:
        try:
            return raw.decode(enc, errors="replace"), enc
        except Exception:
            continue
    return raw.decode("latin-1", errors="replace"), "latin-1"

def _is_markdown_heading(line: str) -> bool:
    return bool(re.match(r"^\s{0,3}#{1,6}\s+\S", line))

def _shorten(s: str, max_chars: int) -> str:
    if len(s) <= max_chars:
        return s
    head = max_chars * 3 // 4
    tail = max_chars - head - 10
    return f"{s[:head]} [...] {s[-tail:]}"

def _plural(n: int, word: str) -> str:
    return f"{n} {word}" + ("" if n == 1 else "s")

def _fmt_dt(ts: float) -> str:
    try:
        return datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return "unknown"

def _try_imports():
    """Lazy import optional heavy deps into module globals."""
    global _pd, _yaml, _openpyxl
    if _pd is None:
        try:
            import pandas as pd  # type: ignore
            _pd = pd
        except Exception:
            _pd = None
    if _yaml is None:
        try:
            import yaml  # type: ignore
            _yaml = yaml
        except Exception:
            _yaml = None
    if _openpyxl is None:
        try:
            import openpyxl  # type: ignore
            _openpyxl = openpyxl
        except Exception:
            _openpyxl = None

def _expand_inputs(files: Union[str, Path, Sequence[Union[str, Path]]]) -> List[Path]:
    """
    Accepts: path/str, list/tuple of paths/str, or globs, or directories.
    Returns a unique, ordered list of matching files (supported extensions only).
    """
    if isinstance(files, (str, Path)):
        files = [files]
    out: List[Path] = []
    seen = set()
    for item in files:  # type: ignore
        p = Path(os.path.expanduser(str(item)))
        if "*" in str(p) or "?" in str(p) or "[" in str(p):
            for fp in sorted(Path().glob(str(p))):
                if fp.is_file() and fp.suffix.lower() in SUPPORTED_EXTS and fp not in seen:
                    out.append(fp); seen.add(fp)
        elif p.is_dir():
            for fp in sorted(p.rglob("*")):
                if fp.is_file() and fp.suffix.lower() in SUPPORTED_EXTS and fp not in seen:
                    out.append(fp); seen.add(fp)
        elif p.is_file():
            if p.suffix.lower() in SUPPORTED_EXTS and p not in seen:
                out.append(p); seen.add(p)
    return out


# ------------------------------ Summarizers -------------------------------- #

@dataclass
class SummarySettings:
    max_output_chars: int = 8000
    sample_rows: int = 5
    top_n_values: int = 5
    read_rows_cap: Optional[int] = None  # for CSV/XLSX; None means let backend decide
    text_preview_bytes: int = 200_000
    json_preview_items: int = 2000   # cap parsing for massive lists
    json_inline_values: int = 5       # how many example values to inline for list/dict
    include_samples: bool = True
    include_stats: bool = True
    include_warnings: bool = True


class _DFProfiler:
    """DataFrame profiling helpers (works best with pandas)."""

    def __init__(self, settings: SummarySettings):
        self.s = settings

    def profile(self, df, name: str = "sheet") -> str:
        lines: List[str] = []
        try:
            n_rows, n_cols = df.shape
        except Exception:
            n_rows = len(df)
            n_cols = getattr(df, "columns", [])
            n_cols = len(n_cols) if n_cols is not None else 0

        lines.append(f"- Shape: {_plural(n_rows, 'row')}, {_plural(n_cols, 'column')}")
        if hasattr(df, "memory_usage"):
            try:
                mem = int(df.memory_usage(deep=True).sum())
                lines.append(f"- In-memory size (approx): {_human_bytes(mem)}")
            except Exception:
                pass

        # Column dictionary
        lines.append("- Columns:")
        for col in list(getattr(df, "columns", []))[:1000]:
            try:
                series = df[col]
            except Exception:
                continue
            col_line = self._profile_col(series, col)
            lines.append(col_line)

        # Missingness hotspots
        if self.s.include_stats:
            try:
                miss = df.isna().mean().sort_values(ascending=False)
                hot = [f"  - {c}: {v*100:.1f}% missing" for c, v in miss.items() if v > 0.2][:20]
                if hot:
                    lines.append("- Missingness hotspots (>20% missing):")
                    lines.extend(hot)
            except Exception:
                pass

        # Sample rows
        if self.s.include_samples:
            try:
                sample = df.head(self.s.sample_rows)
                # Render a compact CSV-style sample (avoids markdown table bloat)
                csv_buf = io.StringIO()
                try:
                    sample.to_csv(csv_buf, index=False)
                    sample_csv = csv_buf.getvalue().strip()
                except Exception:
                    # fallback
                    sample_csv = str(sample.head(self.s.sample_rows))
                lines.append("- Sample (CSV format):")
                for line in sample_csv.splitlines()[: 10 + self.s.sample_rows]:
                    lines.append(f"  {line}")
            except Exception:
                pass

        return "\n".join(lines)

    def _profile_col(self, series, name: str) -> str:
        # dtype & basic counts
        try:
            dtype = str(series.dtype)
        except Exception:
            dtype = "unknown"

        try:
            nunique = int(series.nunique(dropna=True))
        except Exception:
            nunique = -1

        try:
            missing_pct = float(series.isna().mean()) * 100.0
        except Exception:
            missing_pct = float("nan")

        # Basic stats by type (best-effort)
        extras: List[str] = []
        if _pd is not None:
            try:
                if _pd.api.types.is_numeric_dtype(series):
                    desc = series.describe(include="all")
                    mn = desc.get("min", None)
                    mx = desc.get("max", None)
                    mean = desc.get("mean", None)
                    std = desc.get("std", None)
                    if mn is not None and mx is not None:
                        extras.append(f"range=[{mn}, {mx}]")
                    if mean is not None:
                        extras.append(f"mean={mean:.3g}")
                    if std is not None and not math.isnan(std):
                        extras.append(f"std={std:.3g}")
                elif _pd.api.types.is_datetime64_any_dtype(series):
                    try:
                        mn = series.min()
                        mx = series.max()
                        if _pd.notna(mn) and _pd.notna(mx):
                            extras.append(f"temporal=[{mn}, {mx}]")
                    except Exception:
                        pass
                else:
                    # categorical-ish: show top values
                    vc = series.astype(str).value_counts(dropna=True).head(self.s.top_n_values)
                    topv = "; ".join([f"{k} ({int(v)})" for k, v in vc.items()])
                    if topv:
                        extras.append(f"top={topv}")
            except Exception:
                pass

        meta = f"  - {name}: dtype={dtype}"
        if nunique >= 0:
            meta += f", unique={nunique}"
        if not math.isnan(missing_pct):
            meta += f", missing={missing_pct:.1f}%"
        if extras:
            meta += ", " + ", ".join(extras)
        return meta


# ----------------------------- File Handlers ------------------------------- #

class _FileSummarizer:
    def __init__(self, settings: SummarySettings):
        self.s = settings
        self.dfprof = _DFProfiler(settings)

    # ---- CSV ----
    def summarize_csv(self, path: Path) -> str:
        lines = self._header(path)
        if _pd is None:
            lines.append("! pandas not installed; using minimal CSV reader.")
            lines.extend(self._csv_minimal(path))
            return "\n".join(lines)

        read_kwargs = dict(low_memory=False)
        if self.s.read_rows_cap is not None and self.s.read_rows_cap > 0:
            read_kwargs["nrows"] = self.s.read_rows_cap

        try:
            df = _pd.read_csv(path, **read_kwargs)
        except Exception as e:
            lines.append(f"! Failed to read CSV with pandas: {e}")
            lines.extend(self._csv_minimal(path))
            return "\n".join(lines)

        lines.append(self.dfprof.profile(df, name=path.name))
        return "\n".join(lines)

    def _csv_minimal(self, path: Path) -> List[str]:
        # Minimal shape + header + first N rows with stdlib csv
        lines: List[str] = []
        try:
            with open(path, newline="", encoding="utf-8", errors="replace") as f:
                reader = csv.reader(f)
                rows = []
                for i, row in enumerate(reader):
                    if i == 0:
                        header = row
                    rows.append(row)
                    if self.s.read_rows_cap and i >= self.s.read_rows_cap:
                        break
            n_rows = max(len(rows) - 1, 0)
            n_cols = len(header) if rows else 0
            lines.append(f"- Shape (approx): {_plural(n_rows, 'row')}, {_plural(n_cols, 'column')}")
            lines.append("- Header: " + ", ".join(header[:50]))
            if self.s.include_samples:
                lines.append("- Sample (CSV format):")
                for r in rows[: 1 + self.s.sample_rows]:
                    lines.append("  " + ",".join(r))
        except Exception as e:
            lines.append(f"! Could not parse CSV minimally: {e}")
        return lines

    # ---- XLSX ----
    def summarize_xlsx(self, path: Path) -> str:
        lines = self._header(path)
        if _pd is None and _openpyxl is None:
            lines.append("! pandas/openpyxl not installed; minimal XLSX inspection.")
            lines.extend(self._xlsx_minimal(path))
            return "\n".join(lines)

        try:
            # Prefer pandas for profiling; fall back to openpyxl metadata
            if _pd is not None:
                xls = _pd.ExcelFile(path)
                lines.append(f"- Sheets: {', '.join(xls.sheet_names[:20])}")
                for sheet in xls.sheet_names[:20]:
                    kwargs = {}
                    if self.s.read_rows_cap:
                        kwargs["nrows"] = self.s.read_rows_cap
                    try:
                        df = xls.parse(sheet, **kwargs)
                        lines.append(f"\n[Sheet: {sheet}]")
                        lines.append(self.dfprof.profile(df, name=sheet))
                    except Exception as e:
                        lines.append(f"! Failed to read sheet '{sheet}' with pandas: {e}")
            else:
                lines.extend(self._xlsx_minimal(path))
        except Exception as e:
            lines.append(f"! Could not open XLSX: {e}")
        return "\n".join(lines)

    def _xlsx_minimal(self, path: Path) -> List[str]:
        lines: List[str] = []
        try:
            if _openpyxl is None:
                lines.append("! openpyxl not available; cannot introspect sheets.")
                return lines
            wb = _openpyxl.load_workbook(path, read_only=True)
            sheets = wb.sheetnames
            lines.append(f"- Sheets: {', '.join(sheets[:20])}")
            for sheet_name in sheets[:5]:
                sh = wb[sheet_name]
                # crude dimension
                n_rows = sh.max_row
                n_cols = sh.max_column
                lines.append(f"  - {sheet_name}: approx {_plural(n_rows, 'row')}, {_plural(n_cols, 'column')}")
        except Exception as e:
            lines.append(f"! Could not minimally parse XLSX: {e}")
        return lines

    # ---- JSON ----
    def summarize_json(self, path: Path) -> str:
        lines = self._header(path)
        try:
            raw_text, enc = _safe_read_text(path, max_bytes=max(self.s.text_preview_bytes, 200_000))
            data = json.loads(raw_text)
        except Exception as e:
            lines.append(f"! Failed to parse JSON: {e}")
            snippet, enc = _safe_read_text(path, max_bytes=1200)
            lines.append("- Text preview:")
            lines.append("  " + _shorten(snippet.replace("\n", " ")[:1000], 950))
            return "\n".join(lines)

        lines.extend(self._summarize_any_jsonlike(data))
        return "\n".join(lines)

    def _summarize_any_jsonlike(self, data: Any, prefix: str = "") -> List[str]:
        lines: List[str] = []
        if isinstance(data, list):
            lines.append(f"- JSON type: list (len={len(data)})")
            if len(data) == 0:
                return lines
            # If list of dicts, treat as table
            if isinstance(data[0], dict) and _pd is not None:
                try:
                    cap = self.s.json_preview_items
                    df = _pd.DataFrame(data[:cap])
                    lines.append(self.dfprof.profile(df, name="json_list_of_objects"))
                    return lines
                except Exception:
                    pass
            # Otherwise show some element outlines
            lines.append("- Example items:")
            for i, el in enumerate(data[: self.s.json_inline_values]):
                lines.append(f"  - [{i}] {self._inline_value(el)}")
            return lines

        if isinstance(data, dict):
            lines.append("- JSON type: object")
            lines.append("- Top-level keys:")
            for k in list(data.keys())[:100]:
                v = data[k]
                kind = type(v).__name__
                extra = ""
                if isinstance(v, list):
                    extra = f" (list len={len(v)})"
                elif isinstance(v, dict):
                    extra = f" (object with {len(v)} keys)"
                lines.append(f"  - {k}: {kind}{extra}")
            # Show a few key/value examples
            lines.append("- Key examples:")
            for k in list(data.keys())[: self.s.json_inline_values]:
                lines.append(f"  - {k}: {self._inline_value(data[k])}")
            return lines

        # Primitive
        lines.append(f"- JSON primitive: {self._inline_value(data)}")
        return lines

    def _inline_value(self, v: Any) -> str:
        if isinstance(v, (dict, list)):
            s = json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else str(v)
            return _shorten(s, 300)
        if isinstance(v, str):
            return '"' + _shorten(v.replace("\n", " "), 280) + '"'
        return str(v)

    # ---- YAML ----
    def summarize_yaml(self, path: Path) -> str:
        lines = self._header(path)
        if _yaml is None:
            lines.append("! pyyaml not installed; showing text preview only.")
            snippet, enc = _safe_read_text(path, max_bytes=2000)
            lines.append("- Text preview:")
            for ln in _shorten(snippet, 1200).splitlines()[:30]:
                lines.append("  " + ln)
            return "\n".join(lines)
        try:
            text, _enc = _safe_read_text(path, max_bytes=max(self.s.text_preview_bytes, 200_000))
            data = _yaml.safe_load(text)
        except Exception as e:
            lines.append(f"! Failed to parse YAML: {e}")
            snippet, _enc = _safe_read_text(path, max_bytes=1200)
            lines.append("- Text preview:")
            for ln in _shorten(snippet, 1200).splitlines()[:30]:
                lines.append("  " + ln)
            return "\n".join(lines)
        # Reuse JSON summary logic
        lines.extend(self._summarize_any_jsonlike(data))
        return "\n".join(lines)

    # ---- TXT / MD ----
    def summarize_text(self, path: Path) -> str:
        lines = self._header(path)
        text, enc = _safe_read_text(path, max_bytes=self.s.text_preview_bytes)
        # Basic counts
        char_count = len(text)
        line_count = text.count("\n") + 1 if text else 0
        word_count = len(re.findall(r"\S+", text))

        lines.append(f"- Text size: {_plural(line_count, 'line')}, {_plural(word_count, 'word')}, {_plural(char_count, 'char')}")
        # Headings (markdown-ish)
        headings = [ln.strip() for ln in text.splitlines() if _is_markdown_heading(ln)]
        if headings:
            lines.append("- Headings (first 10):")
            for h in headings[:10]:
                lines.append(f"  - {h.strip()}")
        # Bullets (very simple)
        bullets = [ln.strip() for ln in text.splitlines() if re.match(r"^\s*[-*•]\s+\S", ln)]
        if bullets:
            lines.append(f"- Bulleted items (first {self.s.top_n_values}):")
            for b in bullets[: self.s.top_n_values]:
                lines.append(f"  - {b}")

        # Preview
        preview = _shorten(text, 1200)
        if preview:
            lines.append("- Text preview:")
            for ln in preview.splitlines()[:30]:
                lines.append("  " + ln)
        return "\n".join(lines)

    # ---- Shared ----
    def _header(self, path: Path) -> List[str]:
        try:
            size = path.stat().st_size
            mtime = path.stat().st_mtime
        except Exception:
            size = None
            mtime = None
        return [
            f"\n===== File: {path.name} =====",
            f"- Path: {str(path)}",
            f"- Size: {_human_bytes(size)}; Last modified: {_fmt_dt(mtime)}",
        ]


# ------------------------------ Main Tool ---------------------------------- #

class VLMFileSummarizer(BaseTool):
    """
    Summarizes JSON/CSV/TXT/XLSX/YAML files into a compact, VLM-ready textual brief.

    Parameters (apply):
        files: Union[str, Path, Sequence[str|Path]]
            Path(s), directories, or globs (e.g., "data/*.csv"). Directories are scanned recursively.
        max_output_chars: int = 8000
            Maximum characters returned (hard cap). The tool tries to include the most useful info first.
        sample_rows: int = 5
            Number of sample rows to include for tabular data.
        top_n_values: int = 5
            Number of top categorical values to list.
        read_rows_cap: Optional[int] = None
            Limit rows read from CSV/XLSX (useful for very large files). None means no explicit cap.
        include_samples: bool = True
            Whether to include sample rows/preview text.
        include_stats: bool = True
            Whether to include numeric/categorical stats (when available).
        include_warnings: bool = True
            Include notes about degraded parsing or missing libs.

    Returns:
        str — a single VLM-ready textual brief covering all input files.
    """

    default_desc = (
        "Summarize JSON/CSV/TXT/XLSX/YAML files into a compact, VLM-ready text brief "
        "with schema, key stats, examples, and previews."
    )

    # Optional: populate heavy libs once
    def setup(self) -> None:
        _try_imports()

    def apply(
        self,
        files: Union[str, Path, Sequence[Union[str, Path]]],
        max_output_chars: int = 8000,
        sample_rows: int = 5,
        top_n_values: int = 5,
        read_rows_cap: Optional[int] = None,
        include_samples: bool = True,
        include_stats: bool = True,
        include_warnings: bool = True,
    ) -> str:
        # Bind settings
        settings = SummarySettings(
            max_output_chars=max_output_chars,
            sample_rows=sample_rows,
            top_n_values=top_n_values,
            read_rows_cap=read_rows_cap,
            include_samples=include_samples,
            include_stats=include_stats,
            include_warnings=include_warnings,
        )
        summarizer = _FileSummarizer(settings)

        # Expand inputs
        paths = _expand_inputs(files)
        if not paths:
            return (
                "VLM-BRIEF v1\n"
                "! No files matched the input. Supported extensions: "
                + ", ".join(sorted(SUPPORTED_EXTS))
            )

        # Intro
        lines: List[str] = []
        lines.append("VLM-BRIEF v1")
        lines.append(f"Files matched: {len(paths)}")
        lines.append("Guidance: Use this brief as structured context for downstream VLM tasks.")
        lines.append("")

        # Summaries
        for path in paths:
            ext = path.suffix.lower()
            try:
                if ext == ".csv":
                    part = summarizer.summarize_csv(path)
                elif ext == ".xlsx":
                    part = summarizer.summarize_xlsx(path)
                elif ext == ".json":
                    part = summarizer.summarize_json(path)
                elif ext in (".yml", ".yaml"):
                    part = summarizer.summarize_yaml(path)
                elif ext in (".txt", ".md"):
                    part = summarizer.summarize_text(path)
                else:
                    part = f"\n===== File: {path.name} =====\n- Path: {path}\n! Unsupported extension."
            except Exception as e:
                part = f"\n===== File: {path.name} =====\n- Path: {path}\n! Error while summarizing: {e}"
            lines.append(part)

            # Early truncate if needed
            if sum(len(x) for x in lines) > max_output_chars * 1.2:
                lines.append("\n! Output truncated due to max_output_chars limit.")
                break

        # Postprocess: hard cap
        result = "\n".join(lines)
        if len(result) > max_output_chars:
            result = _shorten(result, max_output_chars)

        # Epilogue prompt (helps VLMs use the context)
        epilogue = (
            "\n\n---\n"
            "VLM Consumer Hints:\n"
            "- Treat 'Columns:' and 'Sample (CSV format):' as schema & examples for tabular data.\n"
            "- For JSON/YAML, 'Top-level keys' and 'Key examples' convey structure.\n"
            "- Use 'Missingness hotspots' to guide data cleaning steps.\n"
            "- Ask for specific rows/keys/filters if you need more detail."
        )
        if len(result) + len(epilogue) <= max_output_chars:
            result += epilogue
        return result


# -------------------------- Optional: CLI smoke test ------------------------ #

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Summarize files into a VLM-ready brief.")
    parser.add_argument("files", nargs="+", help="Files/dirs/globs to summarize.")
    parser.add_argument("--max_output_chars", type=int, default=8000)
    parser.add_argument("--sample_rows", type=int, default=5)
    parser.add_argument("--top_n_values", type=int, default=5)
    parser.add_argument("--read_rows_cap", type=int, default=None)
    parser.add_argument("--no_samples", action="store_true")
    parser.add_argument("--no_stats", action="store_true")
    args = parser.parse_args()

    tool = VLMFileSummarizer()
    brief = tool.apply(
        files=args.files,
        max_output_chars=args.max_output_chars,
        sample_rows=args.sample_rows,
        top_n_values=args.top_n_values,
        read_rows_cap=args.read_rows_cap,
        include_samples=not args.no_samples,
        include_stats=not args.no_stats,
    )
    print(brief)
